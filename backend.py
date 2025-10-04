import random
from twilio.rest import Client
import smtplib
from email.message import EmailMessage
import mysql.connector as my
from openpyxl import load_workbook

def sign_in_db(user,password):
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q="select empid,password from employee"
    c.execute(q)
    l1=c.fetchall()
    for i in l1:
        if user==i[0]:
            if password==i[1]:
                return 'employee'
            else:
                return 'invalid password'
    else:    
        q1="select empid,password from employer"
        c.execute(q1)
        l2=c.fetchall()
        a.close()
        for i in l2:
            if user==i[0]:
                if password==i[1]:
                    return 'employer'
                else:
                    return 'invalid password'
        else:
            return 'user name not found'
a=''
def send_now(n):
    try:
        n='+91'+n
        otp = random.randint(100000,999999)
        client= Client(account_sid,auth_token)
        msg = client.messages.create(
            body = f"YOUR OTP IS {otp}",
            from_='+19126893709',
            to = n
        )
        print(otp)
        return otp
    except:
        return False
maindata=[]
def md_append(a):
    maindata.append(a)
def md_pop():
    maindata.pop()
           
def create_username():
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q="select empid from employee"
    c.execute(q)
    b=c.fetchall()
    q1='select empid from employer'
    c.execute(q1)
    b1=c.fetchall()
    b=b+b1
    a.close()
    while True:
        op=random.randint(1000,9999)
        us='#'+maindata[0]+str(op)
        if us in b:
            continue
        else:
            md_append(us)
            print(us)
            return us
def emailsender(e):
    msg=EmailMessage()
    msg['Subject']='verification from job alert'
    msg['From']='jobalert.2022@outlook.com'
    msg['To']=e
    otp=random.randint(1000,9999)
    p=str(otp)
    message='HI kathirvel THIS IS A MAIL VERIFICATION FROM JOB ALERT '+''' 
    +==============================================+
    |                                   |===| '''+p+''' |===|                                     |  
    +==============================================+'''
    msg.set_content(message)
    server=smtplib.SMTP('smtp.office365.com',587)
    server.starttls()
    server.login('jobalert.2022@outlook.com','kathir2906')
    server.send_message(msg)
    return otp
def new_data():
    if maindata[4]=='employee':
        a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
        c=a.cursor()
        dob=maindata[1][6:]+'-'+maindata[1][3:5]+'-'+maindata[1][0:2]
        q="insert into employee values('{}','{}','{}','{}',{},'{}','{}','{}','{}','{}')".format(maindata[10],maindata[0],dob,maindata[2],maindata[3],maindata[5],maindata[6],maindata[7],maindata[8],maindata[9])
        c.execute(q)
        a.commit()
        a.close()
    else:
        a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
        c=a.cursor()
        dob=maindata[1][6:]+'-'+maindata[1][3:5]+'-'+maindata[1][0:2]
        q="insert into employer values('{}','{}','{}','{}',{},'{}','{}','{}','{}','{}','{}')".format(maindata[11],maindata[0],dob,maindata[2],maindata[3],maindata[5],maindata[6],maindata[7],maindata[8],maindata[9],maindata[10])
        c.execute(q)
        a.commit()
        a.close()
def forgat_pass(k):
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q="select mobileno from employee where empid='{}'".format(k)
    c.execute(q)
    b=c.fetchone()
    if b!=None:
        return str(b[0])
    else:
        q="select mobileno from employer where empid='{}'".format(k)
        c.execute(q)
        b=c.fetchone()
        if b!=None:
            return str(b[0])
        else:
            return False
        
def update_pass(id,p):
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    r="select empid from employee"
    c.execute(r)
    m=c.fetchall()
    k="select empid from employer"
    c.execute(k)
    m1=c.fetchall()
    if (id,) in m:
        q1="update employee set password='{}' where empid='{}'".format(p,id)
        c.execute(q1)
    elif (id,) in m1:
        q1="update employer set password='{}' where empid='{}'".format(p,id)
        c.execute(q1)
    a.commit()
    a.close()
def randomchoi():
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q="select * from employee"
    c.execute(q)
    l1=c.fetchall()
    a.close
    return l1
def searchbar(m):
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q="select * from employee where degree like '%{}%'".format(m)
    c.execute(q)
    l1=c.fetchall()
    a.close
    if l1==[]:
        return False
    return l1
        
def createrequest(j,q,s,e):
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q2="select rqcode from request"
    c.execute(q2)
    m=c.fetchall()
    l1=['sk','ab','sr','ak']
    while True:
        op=random.randint(100,999)
        us=random.choice(l1)+str(op)
        if (us,) in m:
            continue
        else:
            break
    q1="insert into request values('{}','{}','{}',{},'{}')".format(us,j,q,s,e)
    c.execute(q1)
    a.commit()
    a.close()

def show_req(m):
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q="select rqcode from request where empid='{}'".format(m)
    c.execute(q)
    n=c.fetchall()
    a.close()
    return n

def deleteone(l):
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q="delete from request where rqcode='{}'".format(l)
    c.execute(q)
    
    a.commit()
    a.close()
def employer_see():
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q="select rqcode,job,minquli,minsalary,cname from employer,request where employer.empid=request.empid order by minsalary"
    c.execute(q)
    k=c.fetchall()
    a.close()
    return k
def searchjob(m):
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q="select rqcode,job,minquli,minsalary,cname from employer,request where employer.empid=request.empid and rqcode='{}' or cname='{}' or job ='{}' or minquli like '%''{}'%'order by minsalary".format(m,m,m,m)
    c.execute(q)
    k=c.fetchall()
    a.close()
    return k
def lastmission(k,i):
    wb=load_workbook('RECUMEFILE.xlsx')
    ws=wb.active
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q="select * from employee where empid='{}'".format(i)
    c.execute(q)
    l1=c.fetchone()
    a.close()
    ws['F6']=str(l1[4])
    ws['J6']=str(l1[0])
    ws['E10']=str(l1[1])
    ws['F11']=str(k[0])
    ws['E12']=str(l1[2])
    ws['F13']=str(k[1])
    ws['F14']=str(k[2])
    ws['F15']=str(l1[3])
    ws['F33']=str(k[7])
    rec=k[4].split('\n')
    rec.pop()
    print(rec)
    for i in range(len(rec)):
        n=str(38+i)
        ws['G'+n]=str(rec[i])
    adds=k[5].split('\n')
    adds.pop()
    for i in range(len(adds)):
        n=str(10+i)
        ws['H'+n]=str(adds[i])
    educ=k[6].split('\n')
    educ.pop()
    for i in range(len(educ)):
        n=str(21+i)
        ws['D'+n]=str(adds[i])
    wb.save('RECUMEFILE.xlsx')

def finalmail(e):
    a=my.connect(host='localhost',user='root',password='Sch00l@',database='jobalert')
    c=a.cursor()
    q="select email from employer,request where employer.empid=request.empid and rqcode='{}'".format(e)
    c.execute(q)
    m=c.fetchone()
    m=m[0]
    a.close()
    msg=EmailMessage()
    msg['Subject']='verification from job alert'
    msg['From']='jobalert.2022@outlook.com'
    msg['To']=m
    
    message='THIS IS A ALERT FROM JOB ALERT A RECUME AS BEEN CAME'
    msg.set_content(message)
    with open('RECUMEFILE.xlsx','rb') as a:
        f_c=a.read()
        f_n=a.name
        msg.add_attachment(f_c,maintype='application',subtype='octet-stream',filename=f_n)
    server=smtplib.SMTP('smtp.office365.com',587)
    server.starttls()
    server.login('jobalert.2022@outlook.com','kathir2906')
    server.send_message(msg)




    
